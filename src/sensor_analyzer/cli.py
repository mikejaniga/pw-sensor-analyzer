"""
Command Line Interface (CLI) module for the sensor analysis project.

Handles retrieving arguments from the user, initializes the analyzer,
controls the file processing, and is responsible for presenting
results in the console and saving them to Excel (.xlsx) format.
"""

import argparse
import logging
import sys
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich import inspect
from rich.console import Console
from rich.logging import RichHandler
from rich_argparse import RichHelpFormatter

from .analyzer import SensorAnalyzer

# Initialize Rich console for nicer output
console = Console()


def setup_logging(debug: bool) -> None:
    """Configures logging. Enables verbose mode with colors if debug=True."""
    level = logging.DEBUG if debug else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(message)s",
        datefmt="[%X]",
        handlers=[RichHandler(rich_tracebacks=True, console=console, show_path=debug)],
    )
    if debug:
        logging.debug("Debug mode enabled")


def save_formatted_excel(df: pd.DataFrame, output_file: str) -> None:
    """
    Saves the DataFrame to an Excel file with formatting applied.
    """
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Analysis Results")

        worksheet = writer.sheets["Analysis Results"]

        # Style definitions
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center")

        # Colors for conditional formatting - bright versions
        fills = {
            "!": PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"
            ),  # Bright Red
            "+": PatternFill(
                start_color="00FF00", end_color="00FF00", fill_type="solid"
            ),  # Bright Green
            "-": PatternFill(
                start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"
            ),  # Darker Grey
        }
        fonts = {
            "!": Font(color="FFFFFF", bold=True),  # White text, bold
            "+": Font(color="000000", bold=True),  # Black text, bold
            "-": Font(color="FFFFFF", bold=True),  # White text, bold
        }

        # Iterate over columns to adjust width and format content
        for col_num, col_name in enumerate(df.columns, 1):  # openpyxl starts from 1
            # Header setup
            header_cell = worksheet.cell(row=1, column=col_num)
            header_cell.font = header_font
            header_cell.alignment = center_alignment

            # Column width calculation
            max_len = (
                max(df[col_name].astype(str).map(len).max(), len(str(col_name))) + 3
            )
            worksheet.column_dimensions[get_column_letter(col_num)].width = max_len

            # Data formatting (alignment and colors)
            if col_name != "Sensor":
                for row_num in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.alignment = center_alignment
                    val = str(cell.value).strip()
                    if val in fills:
                        cell.fill = fills[val]
                        cell.font = fonts[val]
            else:
                # Sensor names alignment (left for readability)
                for row_num in range(2, len(df) + 2):
                    worksheet.cell(row=row_num, column=col_num).alignment = Alignment(
                        horizontal="left"
                    )


def valid_date(s: str) -> datetime:
    """
    Helper function for argparse to validate date format.

    Args:
        s (str): Date string.

    Returns:
        datetime: Date object if format is correct.

    Raises:
        argparse.ArgumentTypeError: If the date format is invalid.
    """
    try:
        return datetime.strptime(s, "%d.%m.%Y")
    except ValueError:
        msg = f"Invalid date format: '{s}'. Expected format is DD.MM.YYYY."
        raise argparse.ArgumentTypeError(msg)


def main() -> None:
    """
    Main function of the console script.

    Initializes the command-line argument parser (argparse), retrieves parameters from the user,
    and then starts the analysis of CSV files using the SensorAnalyzer class.
    Results are presented in the console and/or exported to an Excel file according
    to user requirements.

    Supported arguments:
    -s, --start:  Start date (DD.MM.YYYY) - required
    -e, --end:    End date (DD.MM.YYYY) - required
    -o, --output: Output format ('console', 'excel', 'both') - default 'both'
    -f, --file:   Output Excel filename (optional)
    -d, --dir:    Directory with input CSV files (default 'sample_data')
    --debug:      Debug mode
    """
    parser = argparse.ArgumentParser(
        description="Analyze sensor measurements from CSV files.",
        formatter_class=RichHelpFormatter,
    )
    parser.add_argument(
        "-s", "--start", required=True, type=valid_date, help="Start date (DD.MM.YYYY)"
    )
    parser.add_argument(
        "-e", "--end", required=True, type=valid_date, help="End date (DD.MM.YYYY)"
    )
    parser.add_argument(
        "-o",
        "--output",
        choices=["console", "excel", "both"],
        default="both",
        help="Output format (console, excel, or both (default))",
    )
    parser.add_argument("-f", "--file", help="Output Excel filename (optional)")
    parser.add_argument(
        "-d", "--dir", default="sample_data", help="Directory with CSV files"
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable developer debug mode with rich inspection",
    )

    args = parser.parse_args()

    setup_logging(args.debug)
    log = logging.getLogger(__name__)

    try:
        log.debug(f"Initializing analyzer for directory: {args.dir}")
        analyzer = SensorAnalyzer(data_dir=args.dir)
    except (FileNotFoundError, NotADirectoryError) as e:
        log.error(f"Error: {e}")
        sys.exit(1)

    # Since we used type=valid_date, args.start and args.end are already datetime objects!
    files = analyzer.get_files_in_range(args.start, args.end)

    start_str = args.start.strftime("%d.%m.%Y")
    end_str = args.end.strftime("%d.%m.%Y")

    if not files:
        log.warning(
            f"No CSV files found in range {start_str} to {end_str} in directory '{args.dir}'"
        )
        sys.exit(0)

    log.info(f"Processing {len(files)} files...")
    df = analyzer.process_data(files)

    if df.empty:
        log.warning("No data to process.")
        sys.exit(0)

    # --- RICH INSPECTION BLOCK --------------------
    if args.debug:
        console.rule("[bold cyan]Structures[/bold cyan]")
        log.debug("Analyzer object inspection:")
        # Shows attributes and available methods of the object
        inspect(analyzer, methods=True, docs=False, console=console)

        log.debug("\nResulting DataFrame preview:")
        # log.debug will show formatted matrix
        log.debug(df.head())
        console.rule("[bold cyan]End[/bold cyan]\n")
    # ----------------------------------------------

    if args.output in ["console", "both"]:
        log.info("\nResult Matrix:")
        log.info(df.to_string(index=False))

    if args.output in ["excel", "both"]:
        output_file = args.file if args.file else f"results_{start_str}__{end_str}.xlsx"
        try:
            save_formatted_excel(df, output_file)
            log.info(f"\nResults saved to {output_file}")
        except Exception as e:
            log.error(f"Error saving to Excel: {e}", exc_info=args.debug)
            sys.exit(1)


if __name__ == "__main__":
    main()
